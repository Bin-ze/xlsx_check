import streamlit as st
import pandas as pd
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import PatternFill
import prettytable as pt

def highlight_errors(row, error_columns):
    """
    标记错误列，返回带有样式的单元格
    """
    return ['background-color: yellow' if col in error_columns else '' for col in row.index]

def display_validation_errors(df, validation_errors):

    # 初始化一个字典来保存每行的错误列信息
    error_columns_dict = {}

    for col_name, err_msg, err_indices in validation_errors:
        # 初始化一个空的 DataFrame，用于存放所有错误行
        error_rows = pd.DataFrame()
        st.write(f"{err_msg}:")
        for idx in err_indices:
            idx -= 2
            # 将错误列记录在 error_columns_dict 中
            if idx in error_columns_dict:
                error_columns_dict[idx].append(col_name)
            else:
                error_columns_dict[idx] = [col_name]

            # 将错误行拼接到 error_rows 中，并包含实际的 Excel 行号
            row_with_index = df.loc[[idx]].copy()
            row_with_index.insert(0, 'Excel 行号', idx + 2)  # 在最前面插入一列显示实际的 Excel 行号
            error_rows = pd.concat([error_rows, row_with_index])
            # 将错误行拼接到 error_rows 中
            # error_rows = pd.concat([error_rows, df.loc[[idx]]])

        # 使用 Styler 在展示时标记错误列
        styled_df = error_rows.style.apply(lambda row: highlight_errors(row, error_columns_dict.get(row.name, [])), axis=1)
        st.dataframe(styled_df)

class XlsxCheck:
    def __init__(self, xlsx_root):
        self.xlsx_root = xlsx_root

    @staticmethod
    def read_xlsx(file):
        df = pd.read_excel(file, header=0)
        return df

    @staticmethod
    def validate_data(df):
        errors = []
        year_errors = df[df['年'].isna() | df['年'].astype(str).str.contains(' ')]
        if not year_errors.empty:
            errors.append(("年", f"'年' 列存在错误的行", [i + 2 for i in year_errors.index.tolist()]))
        journal_errors = df[df['刊名'].str.contains(' ')]
        if not journal_errors.empty:
            errors.append(('刊名', "'刊名' 列存在空格的行", [i + 2 for i in journal_errors.index.tolist()]))
        title_errors = df[df['题名'].str.contains(r'".*[^"]$') | df['题名'].str.contains(r'^[^"].*"')]
        if not title_errors.empty:
            errors.append(('题名', "'题名' 列存在未封闭的双引号", [i + 2 for i in title_errors.index.tolist()]))
        unescaped_slash_errors = df[df['题名'].str.contains(r'[^\\]/[^\\]')]
        if not unescaped_slash_errors.empty:
            errors.append(('题名', "'题名' 列存在未转义斜杠的行", [i + 2 for i in unescaped_slash_errors.index.tolist()]))
        issue_errors = df[(df['期'].isna()) | (df['期'].astype(float) > 1000)]
        if not issue_errors.empty:
            errors.append(('期', "'期' 列存在 NaN 或值大于 1000 的行", [i + 2 for i in issue_errors.index.tolist()]))
        return errors

    def mark_errors(self, df, errors):
        red_fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")
        yellow_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
        green_fill = PatternFill(start_color="00FF00", end_color="00FF00", fill_type="solid")
        blue_fill = PatternFill(start_color="0000FF", end_color="0000FF", fill_type="solid")
        wb = Workbook()
        ws = wb.active
        for col_idx, col_name in enumerate(df.columns, 1):
            ws.cell(row=1, column=col_idx, value=col_name)
        for row_idx, row in df.iterrows():
            for col_idx, value in enumerate(row, 1):
                ws.cell(row=row_idx + 2, column=col_idx, value=value)
        for column_name, _, error_rows in errors:
            for idx in error_rows:
                excel_row = idx
                if column_name == '年':
                    ws[f'G{excel_row}'].fill = red_fill
                elif column_name == '刊名':
                    ws[f'B{excel_row}'].fill = yellow_fill
                elif column_name == '题名':
                    ws[f'C{excel_row}'].fill = green_fill
                elif column_name == '期':
                    ws[f'H{excel_row}'].fill = blue_fill
        return wb

    def format_res(self, validation_errors):
        for idx, meta in enumerate(validation_errors):
            i, j, k = meta
            df_new = pd.DataFrame(
                    [
                {
                    "错误类型": f"{j}",
                    "错误行数": ",".join([str(i) for i in k]),
                }]
                )
            
            if idx == 0:
                st.session_state.df = df_new
            else:
                st.session_state.df = pd.concat([df_new, st.session_state.df], axis=0)

        st.dataframe(st.session_state.df, use_container_width=True, hide_index=True)

    def export_xlsx(self, wb, output_path):
        wb.save(output_path)
        return output_path

    def __call__(self, file):
        df = self.read_xlsx(file)
        validation_errors = self.validate_data(df)
        ws = self.mark_errors(df, validation_errors)
        output_path = str(Path(file.name).stem + '_错误检查.xlsx')
        output_path = self.export_xlsx(ws, output_path)
        return output_path, validation_errors, df


def main():
    st.title("XLSX 文件格式检查工具")
    uploaded_file = st.file_uploader("上传 XLSX 文件", type=["xlsx"])

    if uploaded_file is not None:
        xlsx_checker = XlsxCheck(uploaded_file)
        if st.button("执行检查"):
            output_path, validation_errors, df = xlsx_checker(uploaded_file)
            st.success(f"检查完成，结果已保存: {output_path}")
            st.write("错误信息如下所示")
            xlsx_checker.format_res(validation_errors)
            if validation_errors:
                st.write("发现以下错误：")
                display_validation_errors(df, validation_errors)
                # for col_name, err_msg, err_df in validation_errors:
                #     error_rows_df = pd.DataFrame()
                #     st.write(f"{err_msg}:")
                #     # for idx in err_df:
                #         # st.dataframe(df.loc[idx])
                #     error_rows_df = pd.concat([error_rows_df, df.loc[err_df]], axis=0)
                #     st.dataframe(error_rows_df)

            with open(output_path, "rb") as file:
                btn = st.download_button(
                    label="下载标记后的文件",
                    data=file,
                    file_name=output_path,
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                )

if __name__ == "__main__":
    main()
