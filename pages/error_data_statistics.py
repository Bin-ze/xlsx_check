# -*- coding: utf-8 -*-
# Author: Bin-ze
# Email: binze.zero@gmail.com
# Date: 2024/8/15 17:49
# File Name: data_statistics.py

"""
Description of the script or module goes here.

需求一：
数据库中的书籍如果已经入库，那么在期刊目录中标记出来
"""
# Your code starts here
import os
import pandas as pd
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import PatternFill
from difflib import SequenceMatcher
import string
import numpy as np

import streamlit as st

def read_xslx(path):
    df = pd.read_excel(path, header=0)
    return  df

def relaxed_match(title, Book_Title):
    # 去除标点符号
    # titles = title.stem
    # 如果title中包含"24"，直接跳过放宽搜索
    if "24" in title:
        return False, None
    title_no_punct = title.translate(str.maketrans('', '', string.punctuation))
    sim = []
    for idx, book_title in enumerate(Book_Title):
        book_title_no_punct = book_title.translate(str.maketrans('', '', string.punctuation))
        # 计算相似度
        similarity = SequenceMatcher(None, title_no_punct, book_title_no_punct).ratio()
        sim.append(similarity)
    sim = np.array(sim)
    if sim.max() >= 0.85:  # 90%以上的一致性
            return True, sim.argmax()
    return False, None


class data_static:
    def __init__(self, txt_300dpi):
        self.Book = self.decode_txt(txt_300dpi)
        self.Book_Title_journal = [self.check_path(Path(i)).parent.parent.name for i in self.Book]

        self.source_path = str(self.Book[0]).split(self.Book_Title_journal[0])[0]

    def check_path(self, path):
        parts = path.parts
        part2 = list(filter(None, map(lambda s: s.replace("/", ""), parts[-2:])))
        part_new = list(parts[:-2]) + part2
        return Path(os.path.join(*part_new))

    def decode_txt(self, uploaded_file):
        # 直接读取 UploadedFile 对象的内容
        content = uploaded_file.getvalue().decode("utf-8")
        meta = content.splitlines()
        return meta

    # def decode_txt(self, txt):
    #     with open(txt, 'r') as f:
    #         meta = f.readlines()
    #
    #     return  [i.splitlines()[0] for i in meta]

    def relaxed_search_qi(self, Table_contents_df):
        Table_contents_df['命名存在错误'] = False
        for index, row in Table_contents_df.iterrows():
            if not row['已入库']:
                # 生成pdf的上一级目录
                try:
                    journal_name = row['刊名']
                    issue_name = "24" + str(int(row['期'])).zfill(2)  # 假设“期名”由年和期构成
                    if int(row['期']) > 100:
                        issue_name = "24" + str(int(row['期'])).zfill(4)
                    title = row['题名']
                    # 这里的上一级目录是必然存在的，不然应该直接报错
                    up_root = Path(os.path.join(self.source_path, journal_name, issue_name))

                    Book_Title = [Path(i).stem for i in self.Book if str(up_root) in i]
                    if Book_Title == []:
                        continue
                    match, idx = relaxed_match(title, Book_Title)
                    if match:
                        # 构建
                        Table_contents_df.at[index, '命名存在错误'] = True
                except:
                    continue
        return Table_contents_df

    def strict_search(self, Table_contents_df):
        Table_contents_df['已入库'] = False
        mask = []
        for index, row in Table_contents_df.iterrows():
            matched = False
            # 提取刊名、期名和题名
            journal_name = row['刊名']
            try:
                issue_name = "24" + str(int(row['期'])).zfill(2)  # 假设“期名”由年和期构成
                if int(row['期']) > 100:
                    issue_name = "24" + str(int(row['期'])).zfill(4)
            except:
                print(row['期'])
                print(row['题名'])
                print(f"请检查和修正表格 {row['期']} - {row['题名']}")
                mask.append(matched)
                continue

            title = row['题名']
            # 生成文件路径
            file_name = f"{title}.pdf"
            file_path = os.path.join(self.source_path, journal_name, issue_name, file_name)
            try:
                if file_path in self.Book:
                    matched = True
                    mask.append(matched)
                else:
                    mask.append(matched)
            except:
                print(f"表格题名非法中英文引用 {row['期']} - {row['题名']}")
                mask.append(matched)
        Table_contents_df['已入库'] = mask

        return Table_contents_df

    def mark_table(self, Table_contents_df):
        wb = Workbook()
        ws = wb.active

        # 将DataFrame的列标题写入Excel
        for col_idx, col_name in enumerate(Table_contents_df.columns, 1):
            ws.cell(row=1, column=col_idx, value=col_name)

        # 定义红色填充样式
        red_fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")
        green_fill = PatternFill(start_color="00FF00", end_color="00FF00", fill_type="solid")

        # 将DataFrame中的数据写入工作表，并根据'已入库'标记设置颜色
        for row_idx, row in Table_contents_df.iterrows():
            for col_idx, value in enumerate(row, 1):
                cell = ws.cell(row=row_idx + 2, column=col_idx, value=value)
                if row['已入库']:
                    cell.fill = red_fill  # 严格规则匹配成功，标记为红色
                elif row['命名存在错误']:
                    cell.fill = green_fill  # 宽松规则匹配成功，标记为绿色
                    # if row['命名修正']:
                    #     cell.fill = blue_fill  # 宽松规则匹配成功，标记为绿色
        return wb

    def save_res(self, wb, Table_contents, save_path=None):
        print(f"更新完成， 生成表格： {str(Path(Table_contents).parent.joinpath(Path(Table_contents).stem + '_标记.xlsx'))}")
        if save_path is None:
            wb.save(str(Path(Table_contents).parent.joinpath(Path(Table_contents).stem + '_标记.xlsx')))
            return str(Path(Table_contents).parent.joinpath(Path(Table_contents).stem + '_标记.xlsx'))
        else:
            wb.save(save_path)
            return save_path

    def func_1(self, Table_contents, save_path=None):

        Table_contents_df = read_xslx(Table_contents)
        # 非常严格的匹配
        print("\n严格匹配开始\n")
        Table_contents_df = self.strict_search(Table_contents_df)


        print("\n宽松匹配在期目录下查找可能得命名错误\n")
        # 应用宽松规则进行标记，希望可以在错误的期刊中大概找到错误的原因
        Table_contents_df = self.relaxed_search_qi(Table_contents_df)
        # 使用openpyxl创建一个新的工作簿并获取活动工作表

        wb = self.mark_table(Table_contents_df)

        output_path = self.save_res(wb, Table_contents.name, save_path)

        return Table_contents_df, wb, output_path

def main():
    st.set_page_config(page_title="期刊目录与数据库统计工具", page_icon="📈")
    st.markdown("# 期刊目录与数据库统计工具")
    st.sidebar.header("期刊目录与数据库统计工具")
    st.write(
        """该工具统计表格与数据库中文件的对照关系，并标记"""
    )
    # 上传表格文件
    Table_contents = st.file_uploader("上传期刊目录 Excel 文件", type=["xlsx"])
    txt_300dpi = st.file_uploader("上传300 dpi txt", type=["txt"])

    if Table_contents and txt_300dpi:
        static = data_static(txt_300dpi)
        if st.button("执行统计"):
            with st.spinner('统计中，请稍候...'):
                # 初始化 data_static 实例
                # 调用 func_1 进行统计
                result_df, wb, output_path = static.func_1(Table_contents)

                # 计算匹配统计数据
                total_entries = len(result_df)
                strict_matches = result_df['已入库'].sum()
                relaxed_matches = result_df['命名存在错误'].sum()
                unmatched = total_entries - (strict_matches + relaxed_matches)

                # 可视化匹配统计数据
                col1, col2, col3 = st.columns(3)
                col2.metric("total_entries", f"{total_entries}")
                col1.metric("strict_matches", f"{strict_matches}", f"{(round(strict_matches / total_entries, 3) * 100):.3f}")
                col2.metric("relaxed_matches", f"{relaxed_matches}", f"{(round(relaxed_matches / total_entries, 3) * 100):.3f}")
                col3.metric("unmatched", f"{unmatched}", f"{(round(unmatched / total_entries, 3) * 100):.3f}")

                # 显示统计结果的详细 DataFrame
                # st.write("匹配结果详细数据：")
                # st.dataframe(result_df)
                # 展示“命名存在错误”的行
                # 展示“命名存在错误”的行
                st.write("命名存在错误的行：")
                relaxed_error_rows = result_df[result_df['命名存在错误'] == 1]
                relaxed_error_rows.index = relaxed_error_rows.index + 2  # 调整行号
                st.dataframe(relaxed_error_rows)

                # 展示“unmatched”的行
                st.write("未匹配的行：")
                unmatched_rows = result_df[(result_df['已入库'] == 0) & (result_df['命名存在错误'] == 0)]
                unmatched_rows.index = unmatched_rows.index + 2  # 调整行号
                st.dataframe(unmatched_rows)

                st.markdown("## 请从此处下载标记之后的xlsx文件")
                # 你也可以让用户下载标记后的文件
                with open(output_path, "rb") as file:
                    btn = st.download_button(
                        label="下载标记后的文件",
                        data=file,
                        file_name=output_path,
                        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                    )

if __name__ == '__main__':
    main()

# if __name__ == '__main__':
#     Table_contents, sourse_path = "/Users/binze/Desktop/work_code/statistics_pcf/第一批数据/第一批期刊目录_标记_v3_test.xlsx", "/Users/binze/Desktop/work_code/statistics_pcf/第一批数据/300dpi.txt"
#
#     static = data_static(sourse_path)
#     # fun1数据库中的书籍如果已经入库，那么在期刊目录中标记出来
#     # static.func_1(Table_contents=Table_contents)
#
#     result_df, wb, output_path = static.func_1(Table_contents)
#
#     # 计算匹配统计数据
#     total_entries = len(result_df)
#     strict_matches = result_df['已入库'].sum()
#     relaxed_matches = result_df['命名存在错误'].sum()
#     unmatched = total_entries - (strict_matches + relaxed_matches)